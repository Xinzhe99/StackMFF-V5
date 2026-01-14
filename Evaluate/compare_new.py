# -*- coding: utf-8 -*-
# @Author  : XinZhe Xie
# @University  : ZheJiang University

# Usage example:
# python compare_new.py \
#     --base_path /path/to/base \
#     --methods CVT DWT DCT DTCWT NSCT IFCNN-MAX U2Fusion SDNet MFF-GAN SwinFusion MUFusion SwinMFF DDBFusion CCSR-Net MCCSR-Net "Zerene Stacker - DMap" "Zerene Stacker - PMax" "Helicon Focus 8 - A" "Helicon Focus 8 - B" "Helicon Focus 8 - C" StackMFF "StackMFF V2" "StackMFF V3" "StackMFF V4" "StackMFF V5" \
#     --datasets "Mobile Depth" Middlebury \
#     --metrics BRISQUE PIQE LPIPS EN AG SF \
#     --output_dir ./evaluation_outputs

import numpy as np
import pandas as pd
import cv2
import glob
from tqdm import tqdm
import warnings
import argparse
import torch
import os
os.environ['HF_ENDPOINT'] = 'https://hf-mirror.com'
import pyiqa
from datetime import datetime

from skimage.measure import shannon_entropy
from skimage.metrics import peak_signal_noise_ratio, structural_similarity
import numpy as np

warnings.filterwarnings('ignore')

def parse_args():
    parser = argparse.ArgumentParser(description='Compare different image fusion methods by calculating evaluation metrics')
    parser.add_argument('--base_path', type=str, default=r'/home/ot/Students/xxz/datasets/results_of_different_methods',
                        help='Base path containing the result images for different methods')
    # parser.add_argument('--methods', type=str, nargs='+', default=['StackMFF V5'],
    #                     help='List of methods to compare')
    parser.add_argument('--methods', type=str, nargs='+', default=['CVT', 
                                                                    'DWT', 
                                                                    'DCT', 
                                                                    'DTCWT', 
                                                                    'NSCT', 
                                                                    'IFCNN-MAX', 
                                                                    'U2Fusion', 
                                                                    'SDNet', 
                                                                    'MFF-GAN', 
                                                                    'SwinFusion', 
                                                                    'MUFusion', 
                                                                    'SwinMFF', 
                                                                    'DDBFusion', 
                                                                    'CCSR-Net', 
                                                                    'MCCSR-Net', 
                                                                    'Zerene Stacker - DMap', 
                                                                    'Zerene Stacker - PMax', 
                                                                    'Helicon Focus 8 - A', 
                                                                    'Helicon Focus 8 - B', 
                                                                    'Helicon Focus 8 - C', 
                                                                    'StackMFF', 
                                                                    'StackMFF V2', 
                                                                    'StackMFF V3', 
                                                                    'StackMFF V4',
                                                                    'StackMFF V5'])
    parser.add_argument('--datasets', type=str, nargs='+', 
                        default=['Mobile Depth','Middlebury','FlyingThings3D','Road-MF'],
                        help='List of datasets to evaluate')
    # parser.add_argument('--metrics', type=str, nargs='+', default=['BRISQUE', 'PIQE', 'NIQE', 'MUSIQ', 'LPIPS', 'VIF', 'DISTS', 'PSNR', 'SSIM', 'MS_SSIM', 'MSE', 'MAE', 'RMSE', 'logRMS', 'EN', 'AG', 'SF'],
    #                     help='List of metrics to calculate')
    parser.add_argument('--metrics', type=str, nargs='+', default=['SSIM','PSNR'],
                        help='List of metrics to calculate')
    parser.add_argument('--enable_registration', default='false',
                        help='Enable image registration')
    parser.add_argument('--output_dir', type=str, default='./evaluation_outputs',
                        help='Directory to save output results')
    return parser.parse_args()

# Initialize models
brisque_model = None
piqe_model = None
lpips_model = None
vif_model = None
dists_model = None
niqe_model = None
musiq_model = None


def get_brisque_model():
    global brisque_model
    if brisque_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        brisque_model = pyiqa.create_metric('brisque', device=device)
    return brisque_model

def get_piqe_model():
    global piqe_model
    if piqe_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        piqe_model = pyiqa.create_metric('piqe', device=device)
    return piqe_model

def get_lpips_model():
    global lpips_model
    if lpips_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        lpips_model = pyiqa.create_metric('lpips', device=device)
    return lpips_model


def get_vif_model():
    global vif_model
    if vif_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        vif_model = pyiqa.create_metric('vif', device=device)
    return vif_model


def get_dists_model():
    global dists_model
    if dists_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        dists_model = pyiqa.create_metric('dists', device=device)
    return dists_model


def get_niqe_model():
    global niqe_model
    if niqe_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        niqe_model = pyiqa.create_metric('niqe', device=device)
    return niqe_model


def get_fid_model():
    global fid_model
    if fid_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        fid_model = pyiqa.create_metric('fid', device=device)
    return fid_model


def get_musiq_model():
    global musiq_model
    if musiq_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        musiq_model = pyiqa.create_metric('musiq', device=device)
    return musiq_model


def get_ms_ssim_model():
    global ms_ssim_model
    if ms_ssim_model is None:
        device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
        ms_ssim_model = pyiqa.create_metric('ms_ssim', device=device)
    return ms_ssim_model

def calculate_pyiqa_metric(img, metric_model):
    if len(img.shape) == 2:
        img_rgb = cv2.cvtColor(img, cv2.COLOR_GRAY2RGB)
    else:
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    img_rgb = img_rgb.astype(np.float32) / 255.0
    img_tensor = torch.from_numpy(img_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img_tensor = img_tensor.cuda()
    with torch.no_grad():
        score = metric_model(img_tensor)
    return score.item()


def calculate_fid(img1, img2, fid_model):
    # FID is typically calculated between two sets of images, but for single image comparison
    # we'll use it similar to other reference-based metrics
    if len(img1.shape) == 2:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_GRAY2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_GRAY2RGB)
    else:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_BGR2RGB)
    img1_rgb = img1_rgb.astype(np.float32) / 255.0
    img2_rgb = img2_rgb.astype(np.float32) / 255.0
    img1_tensor = torch.from_numpy(img1_rgb).permute(2, 0, 1).unsqueeze(0)
    img2_tensor = torch.from_numpy(img2_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img1_tensor = img1_tensor.cuda()
        img2_tensor = img2_tensor.cuda()
    with torch.no_grad():
        score = fid_model(img1_tensor, img2_tensor)
    return score.item()


def calculate_ms_ssim(img1, img2, ms_ssim_model):
    if len(img1.shape) == 2:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_GRAY2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_GRAY2RGB)
    else:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_BGR2RGB)
    img1_rgb = img1_rgb.astype(np.float32) / 255.0
    img2_rgb = img2_rgb.astype(np.float32) / 255.0
    img1_tensor = torch.from_numpy(img1_rgb).permute(2, 0, 1).unsqueeze(0)
    img2_tensor = torch.from_numpy(img2_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img1_tensor = img1_tensor.cuda()
        img2_tensor = img2_tensor.cuda()
    with torch.no_grad():
        score = ms_ssim_model(img1_tensor, img2_tensor)
    return score.item()

# LPIPS
def calculate_lpips(img1, img2, lpips_model):
    if len(img1.shape) == 2:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_GRAY2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_GRAY2RGB)
    else:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_BGR2RGB)
    img1_rgb = img1_rgb.astype(np.float32) / 255.0
    img2_rgb = img2_rgb.astype(np.float32) / 255.0
    img1_tensor = torch.from_numpy(img1_rgb).permute(2, 0, 1).unsqueeze(0)
    img2_tensor = torch.from_numpy(img2_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img1_tensor = img1_tensor.cuda()
        img2_tensor = img2_tensor.cuda()
    with torch.no_grad():
        score = lpips_model(img1_tensor, img2_tensor)
    return score.item()


def calculate_vif(img1, img2, vif_model):
    if len(img1.shape) == 2:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_GRAY2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_GRAY2RGB)
    else:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_BGR2RGB)
    img1_rgb = img1_rgb.astype(np.float32) / 255.0
    img2_rgb = img2_rgb.astype(np.float32) / 255.0
    img1_tensor = torch.from_numpy(img1_rgb).permute(2, 0, 1).unsqueeze(0)
    img2_tensor = torch.from_numpy(img2_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img1_tensor = img1_tensor.cuda()
        img2_tensor = img2_tensor.cuda()
    with torch.no_grad():
        score = vif_model(img1_tensor, img2_tensor)
    return score.item()


def calculate_dists(img1, img2, dists_model):
    if len(img1.shape) == 2:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_GRAY2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_GRAY2RGB)
    else:
        img1_rgb = cv2.cvtColor(img1, cv2.COLOR_BGR2RGB)
        img2_rgb = cv2.cvtColor(img2, cv2.COLOR_BGR2RGB)
    img1_rgb = img1_rgb.astype(np.float32) / 255.0
    img2_rgb = img2_rgb.astype(np.float32) / 255.0
    img1_tensor = torch.from_numpy(img1_rgb).permute(2, 0, 1).unsqueeze(0)
    img2_tensor = torch.from_numpy(img2_rgb).permute(2, 0, 1).unsqueeze(0)
    if torch.cuda.is_available():
        img1_tensor = img1_tensor.cuda()
        img2_tensor = img2_tensor.cuda()
    with torch.no_grad():
        score = dists_model(img1_tensor, img2_tensor)
    return score.item()


def calculate_mse(img1, img2):
    # Convert to float and ensure same dimensions
    if len(img1.shape) == 2:
        img1 = img1.astype(np.float64) / 255.0  # Normalize to [0, 1]
        img2 = img2.astype(np.float64) / 255.0
    else:
        img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
        img2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
    
    # Resize if dimensions don't match
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    mse = np.mean((img1 - img2) ** 2)
    return mse

def calculate_mae(img1, img2):
    # Convert to float and ensure same dimensions
    if len(img1.shape) == 2:
        img1 = img1.astype(np.float64) / 255.0  # Normalize to [0, 1]
        img2 = img2.astype(np.float64) / 255.0
    else:
        img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
        img2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
    
    # Resize if dimensions don't match
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    mae = np.mean(np.abs(img1 - img2))
    return mae

def calculate_rmse(img1, img2):
    # Convert to float and ensure same dimensions
    if len(img1.shape) == 2:
        img1 = img1.astype(np.float64) / 255.0  # Normalize to [0, 1]
        img2 = img2.astype(np.float64) / 255.0
    else:
        img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
        img2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
    
    # Resize if dimensions don't match
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    mse = np.mean((img1 - img2) ** 2)
    rmse = np.sqrt(mse)
    return rmse

def calculate_log_rms(img1, img2):
    # Convert to float and ensure same dimensions
    if len(img1.shape) == 2:
        img1 = img1.astype(np.float64) / 255.0  # Normalize to [0, 1]
        img2 = img2.astype(np.float64) / 255.0
    else:
        img1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
        img2 = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY).astype(np.float64) / 255.0
    
    # Resize if dimensions don't match
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    # Calculate RMS error
    mse = np.mean((img1 - img2) ** 2)
    rmse = np.sqrt(mse)
    
    # Calculate log RMS error (log10 of RMSE + 1 to avoid log(0) and provide meaningful log scale)
    log_rms = np.log10(rmse + 1.0)
    return log_rms


def calculate_psnr(img1, img2):
    # Ensure images have the same dimensions
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    # Method compare.py logic:
    # Use skimage default behavior (which infers data_range from dtype)
    # Typically this means data_range=255 for uint8 images
    
    psnr_value = peak_signal_noise_ratio(img1, img2)
    return psnr_value

def calculate_ssim(img1, img2):
    # Ensure images have the same dimensions
    if img1.shape != img2.shape:
        img2 = cv2.resize(img2, (img1.shape[1], img1.shape[0]))
    
    # Convert to grayscale if images are multi-channel
    if len(img1.shape) == 3:
        img1_gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
        img2_gray = cv2.cvtColor(img2, cv2.COLOR_BGR2GRAY)
    else:
        img1_gray = img1
        img2_gray = img2
    
    # Method compare.py logic:
    # "multichannel=False" is default behavior for 2D inputs in skimage < 0.19, 
    # for newer versions it's "channel_axis=None".
    # Importantly, Method compare.py does NOT specify data_range explicitly
    # and lets skimage infer it (usually 255 for uint8).
    
    try:
        # Try new API first (scikit-image >= 0.19)
        ssim_value = structural_similarity(img1_gray, img2_gray, channel_axis=None)
    except TypeError:
        # Fallback for older versions
        ssim_value = structural_similarity(img1_gray, img2_gray, multichannel=False)
        
    return ssim_value

# EN
def calculate_entropy(img):
    if len(img.shape) == 3:
        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        img_gray = img
    return shannon_entropy(img_gray)

# AG
def calculate_average_gradient(img):
    if len(img.shape) == 3:
        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        img_gray = img
    gx = cv2.Sobel(img_gray, cv2.CV_64F,1,0,ksize=3)
    gy = cv2.Sobel(img_gray, cv2.CV_64F,0,1,ksize=3)
    ag = np.mean(np.sqrt(gx**2 + gy**2))
    return ag

# SF
def calculate_spatial_frequency(img):
    if len(img.shape) == 3:
        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        img_gray = img
    rf = np.std(np.diff(img_gray, axis=0))
    cf = np.std(np.diff(img_gray, axis=1))
    sf = np.sqrt(rf**2 + cf**2)
    return sf

def create_output_dir(output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

def get_images_dict(folder, supported_formats=['jpg', 'jpeg', 'png']):
    images_dict = {}
    for fmt in supported_formats:
        for file_path in glob.glob(os.path.join(folder, f"*.{fmt}")):
            name_without_ext = os.path.splitext(os.path.basename(file_path))[0]
            images_dict[name_without_ext] = file_path
        for file_path in glob.glob(os.path.join(folder, f"*.{fmt.upper()}")):
            name_without_ext = os.path.splitext(os.path.basename(file_path))[0]
            images_dict[name_without_ext] = file_path
    return images_dict

def main():
    args = parse_args()
    create_output_dir(args.output_dir)

    print("\n" + "="*60)
    print("METRICS EVALUATION DIRECTION")
    print("="*60)
    device = torch.device("cuda") if torch.cuda.is_available() else torch.device("cpu")
    for metric in args.metrics:
        if metric in ['BRISQUE','PIQE','EN','AG','SF']:
            direction = "Lower is better ↓" if metric in ['BRISQUE','PIQE'] else "Higher is better ↑"
        elif metric in ['LPIPS','VIF','DISTS','MSE','MAE','RMSE','logRMS']:
            direction = "Lower is better ↓" if metric in ['LPIPS','DISTS','MSE','MAE','RMSE','logRMS'] else "Higher is better ↑"
        elif metric in ['PSNR','SSIM']:
            direction = "Higher is better ↑"
        elif metric in ['MS_SSIM']:
            direction = "Higher is better ↑"
        elif metric in ['NIQE']:
            direction = "Lower is better ↓"
        elif metric in ['MUSIQ']:
            direction = "Higher is better ↑"
        else:
            direction = "Unknown"
        print(f"{metric:12s}: {direction}")
    print("="*60 + "\n")

    all_datasets_results = {}
    all_methods_results = {method:{} for method in args.methods}

    for dataset in args.datasets:
        print('Testing:', dataset)
        method_results = {}
        df = pd.DataFrame(columns=(['Method'] + args.metrics))

        for method in args.methods:
            method_images_dict = get_images_dict(os.path.join(args.base_path, method, dataset))
            if not method_images_dict:
                print(f"Warning: No result images found for method {method} in dataset {dataset}")
                continue
            print(f"Found {len(method_images_dict)} result images for method {method}")

            metric_dict = {metric:0 for metric in args.metrics}
            valid_image_count = 0

            for img_name in tqdm(method_images_dict.keys()):
                img_path = method_images_dict[img_name]
                img = cv2.imread(img_path)
                if img is None:
                    print(f"Warning: Could not read image {img_path}")
                    continue
                valid_image_count += 1

                for metric in args.metrics:
                    try:
                        if metric=='BRISQUE':
                            value = calculate_pyiqa_metric(img, get_brisque_model())
                        elif metric=='PIQE':
                            value = calculate_pyiqa_metric(img, get_piqe_model())
                        elif metric=='LPIPS':
                            # Change "reference_images" to "Ground Truth"
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match (LPIPS requires same size)
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                value = calculate_lpips(ref_img, img, get_lpips_model())
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        elif metric=='VIF':
                            # VIF requires reference image like LPIPS
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                value = calculate_vif(ref_img, img, get_vif_model())
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        elif metric=='DISTS':
                            # DISTS requires reference image like LPIPS
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                value = calculate_dists(ref_img, img, get_dists_model())
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        elif metric in ['MSE', 'MAE', 'RMSE', 'logRMS']:
                            # These metrics require reference image
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                if metric == 'MSE':
                                    value = calculate_mse(ref_img, img)
                                elif metric == 'MAE':
                                    value = calculate_mae(ref_img, img)
                                elif metric == 'RMSE':
                                    value = calculate_rmse(ref_img, img)
                                elif metric == 'logRMS':
                                    value = calculate_log_rms(ref_img, img)
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        elif metric in ['PSNR', 'SSIM']:
                            # These metrics require reference image
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                if metric == 'PSNR':
                                    value = calculate_psnr(ref_img, img)
                                elif metric == 'SSIM':
                                    value = calculate_ssim(ref_img, img)
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        elif metric=='EN':
                            value = calculate_entropy(img)
                        elif metric=='AG':
                            value = calculate_average_gradient(img)
                        elif metric=='SF':
                            value = calculate_spatial_frequency(img)
                        elif metric=='NIQE':
                            value = calculate_pyiqa_metric(img, get_niqe_model())
                        elif metric=='MUSIQ':
                            value = calculate_pyiqa_metric(img, get_musiq_model())
                        elif metric=='MS_SSIM':
                            # MS-SSIM requires reference image
                            gt_base_dir = os.path.join(args.base_path, "Ground Truth", dataset)
                            
                            # Support multiple suffixes
                            possible_exts = ['.png', '.jpg', '.jpeg', '.bmp', '.tif']
                            ref_img_path = None
                            
                            # Try to find the file
                            for ext in possible_exts:
                                temp_path = os.path.join(gt_base_dir, f"{img_name}{ext}")
                                if os.path.exists(temp_path):
                                    ref_img_path = temp_path
                                    break
                            
                            if ref_img_path:
                                ref_img = cv2.imread(ref_img_path)
                                # Resize reference if dimensions don't match
                                if ref_img.shape != img.shape:
                                    ref_img = cv2.resize(ref_img, (img.shape[1], img.shape[0]))
                                value = calculate_ms_ssim(ref_img, img, get_ms_ssim_model())
                            else:
                                # Print debug info only for the first failed image to avoid spamming
                                if valid_image_count == 1:
                                    print(f"Warning: GT not found in {gt_base_dir} for {img_name}")
                                value = np.nan
                        else:
                            continue
                        metric_dict[metric] += value
                    except Exception as e:
                        print(f"Error calculating {metric} for {img_name}: {str(e)}")
                        metric_dict[metric] += 0

            if valid_image_count > 0:
                for metric in metric_dict:
                    metric_dict[metric] /= valid_image_count
                    metric_dict[metric] = round(metric_dict[metric],4)
            method_results[method] = metric_dict
            all_methods_results[method][dataset] = metric_dict

            # Write to dataframe
            new_row = {'Method': method}
            new_row.update(metric_dict)
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        all_datasets_results[dataset] = df
        # Create a copy of the dataframe with metric direction indicators
        df_display = df.copy()
        
        # Update column names to include direction indicators
        new_columns = []
        for col in df_display.columns:
            if col == 'Method':
                new_columns.append(col)
            elif col in ['BRISQUE', 'PIQE', 'LPIPS', 'DISTS', 'MSE', 'MAE', 'RMSE', 'logRMS', 'NIQE']:
                new_columns.append(f"{col} ↓")
            else:  # ['VIF', 'PSNR', 'SSIM', 'MS_SSIM', 'EN', 'AG', 'SF', 'MUSIQ']
                new_columns.append(f"{col} ↑")
        
        df_display.columns = new_columns
        print(df_display.to_string(index=False))

    # Save merged results
    if len(args.datasets) > 1:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        data_rows = []
        for method in args.methods:
            row = [method]
            for dataset in args.datasets:
                if dataset in all_methods_results[method]:
                    for metric in args.metrics:
                        row.append(all_methods_results[method][dataset].get(metric,''))
                else:
                    row.extend(['']*len(args.metrics))
            data_rows.append(row)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f'compare_result_merged_{timestamp}.xlsx'
        output_filepath = os.path.join(args.output_dir, output_filename)

        wb = Workbook()
        ws = wb.active
        ws.title = 'Comparison Results'

        # Header row 1: dataset names
        ws.cell(row=1, column=1,value='Datasets')
        col_idx=2
        for dataset in args.datasets:
            start_col = col_idx
            end_col = col_idx + len(args.metrics) - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=dataset)
            cell.alignment = Alignment(horizontal='center',vertical='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3',end_color='D3D3D3',fill_type='solid')
            col_idx=end_col+1

        # Header row 2: metric names
        ws.cell(row=2, column=1,value='Method')
        col_idx=2
        for dataset in args.datasets:
            for metric in args.metrics:
                metric_name = f"{metric} ↓" if metric in ['BRISQUE','PIQE','LPIPS','DISTS','MSE','MAE','RMSE','logRMS','NIQE'] else f"{metric} ↑"
                cell = ws.cell(row=2, column=col_idx,value=metric_name)
                cell.alignment = Alignment(horizontal='center',vertical='center')
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E8E8E8',end_color='E8E8E8',fill_type='solid')
                col_idx+=1

        # Write data rows
        for row_idx,row_data in enumerate(data_rows,start=3):
            for col_idx,value in enumerate(row_data,start=1):
                cell = ws.cell(row=row_idx,column=col_idx,value=value)
                cell.alignment = Alignment(horizontal='center',vertical='center') if col_idx>1 else Alignment(horizontal='left',vertical='center')

        wb.save(output_filepath)
        print(f"\nMerged results saved to {output_filepath}")

    print("\nEND OF EVALUATION")

if __name__=='__main__':
    main()
